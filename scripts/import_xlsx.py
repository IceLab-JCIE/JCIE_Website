#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


class ImportErrorExit(RuntimeError):
    pass


def _strip(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_int(v: Any, *, ctx: str) -> Optional[int]:
    s = _strip(v)
    if not s:
        return None
    try:
        return int(s)
    except Exception as e:
        raise ImportErrorExit(f"{ctx}: expected integer, got {s!r}") from e


def _to_bool(v: Any) -> bool:
    s = _strip(v).lower()
    if s in ("", "0", "false", "no", "n", "off"):
        return False
    if s in ("1", "true", "yes", "y", "on"):
        return True
    # If user typed something else, treat as false but avoid silent weirdness.
    raise ImportErrorExit(f"is_outstanding: expected TRUE/FALSE or 1/0, got {v!r}")


def _split_semicolon(v: Any) -> List[str]:
    s = _strip(v)
    if not s:
        return []
    return [p.strip() for p in s.split(";") if p.strip()]

def _split_domain(v: Any) -> List[str]:
    s = _strip(v).upper()
    if not s:
        return []
    # Accept common separators: "EDA;LCA", "EDA/LCA", "EDA,LCA", "EDA|LCA"
    for ch in ["/", ",", "|"]:
        s = s.replace(ch, ";")
    parts = [p.strip() for p in s.split(";") if p.strip()]
    out: List[str] = []
    seen: set[str] = set()
    for p in parts:
        if p in ("EDA", "LCA") and p not in seen:
            out.append(p)
            seen.add(p)
        elif p:
            raise ImportErrorExit(f"domain: invalid token {p!r} (expected EDA/LCA)")
    return out

def _norm_name(s: str) -> str:
    s = _strip(s).lower()
    for ch in [".", ",", "’", "'", "\"", "–", "-", "鈥?", "�"]:
        s = s.replace(ch, " ")
    s = " ".join(s.split())
    return s


def _yaml_quote(s: str) -> str:
    return json.dumps(s, ensure_ascii=False)


def _yaml_scalar(v: Any) -> str:
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    return _yaml_quote(str(v))


def _yaml_write(obj: Any, *, indent: int = 0) -> str:
    pad = " " * indent
    if isinstance(obj, dict):
        lines: List[str] = []
        for k, v in obj.items():
            if isinstance(v, (dict, list)):
                lines.append(f"{pad}{k}:")
                lines.append(_yaml_write(v, indent=indent + 2).rstrip("\n"))
            else:
                if isinstance(v, str) and "\n" in v:
                    lines.append(f"{pad}{k}: |")
                    for ln in v.splitlines():
                        lines.append(" " * (indent + 2) + ln)
                else:
                    lines.append(f"{pad}{k}: {_yaml_scalar(v)}")
        return "\n".join(lines) + "\n"
    if isinstance(obj, list):
        lines = []
        for item in obj:
            if isinstance(item, dict):
                lines.append(f"{pad}-")
                lines.append(_yaml_write(item, indent=indent + 2).rstrip("\n"))
            elif isinstance(item, list):
                lines.append(f"{pad}-")
                lines.append(_yaml_write(item, indent=indent + 2).rstrip("\n"))
            else:
                if isinstance(item, str) and "\n" in item:
                    lines.append(f"{pad}- |")
                    for ln in item.splitlines():
                        lines.append(" " * (indent + 2) + ln)
                else:
                    lines.append(f"{pad}- {_yaml_scalar(item)}")
        return "\n".join(lines) + "\n"
    return f"{pad}{_yaml_scalar(obj)}\n"


def _write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8", newline="\n")


def _load_workbook(xlsx: Path):
    try:
        import openpyxl  # type: ignore
    except Exception as e:  # pragma: no cover
        raise ImportErrorExit(
            "XLSX import requires openpyxl. Install with: python -m pip install openpyxl"
        ) from e
    return openpyxl.load_workbook(xlsx, read_only=True, data_only=True)


def _read_sheet_rows(ws) -> Tuple[List[str], List[Dict[str, Any]]]:
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return [], []
    headers = [_strip(h).strip() for h in header_row]
    headers_lc = [h.lower() for h in headers]
    out: List[Dict[str, Any]] = []
    for r in rows:
        if r is None:
            continue
        d: Dict[str, Any] = {}
        any_val = False
        for i, h in enumerate(headers_lc):
            val = r[i] if i < len(r) else None
            if val is not None and _strip(val) != "":
                any_val = True
            d[h] = val
        if any_val:
            out.append(d)
    return headers_lc, out


def _require_columns(headers: List[str], required: Iterable[str], *, sheet: str) -> None:
    missing = [c for c in required if c.lower() not in headers]
    if missing:
        raise ImportErrorExit(f"Sheet {sheet!r} missing required columns: {', '.join(missing)}")


@dataclass
class PeopleIndex:
    id_to_name_en: Dict[str, str]
    name_norm_to_id: Dict[str, str]


def _parse_people(rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], PeopleIndex]:
    seen: set[str] = set()
    people: List[Dict[str, Any]] = []
    id_to_name_en: Dict[str, str] = {}
    name_norm_to_id: Dict[str, str] = {}

    for i, r in enumerate(rows, start=2):
        pid = _strip(r.get("id"))
        if not pid:
            raise ImportErrorExit(f"people row {i}: missing id")
        if pid in seen:
            raise ImportErrorExit(f"people: duplicate id {pid!r}")
        seen.add(pid)

        role = _strip(r.get("role"))
        if role not in ("mentor", "member", "alumni"):
            raise ImportErrorExit(f"people row {i} (id={pid}): invalid role {role!r}")

        name_en = _strip(r.get("name_en"))
        name_zh = _strip(r.get("name_zh"))
        if not name_en or not name_zh:
            raise ImportErrorExit(f"people row {i} (id={pid}): name_en and name_zh are required")

        join_year = _to_int(r.get("join_year"), ctx=f"people row {i} (id={pid}) join_year")
        domain_list = _split_domain(r.get("domain"))
        photo = _strip(r.get("photo"))

        title_en = _strip(r.get("title_en"))
        title_zh = _strip(r.get("title_zh"))
        bio_en = _strip(r.get("bio_en"))
        bio_zh = _strip(r.get("bio_zh"))
        destination_en = _strip(r.get("destination_en"))
        destination_zh = _strip(r.get("destination_zh"))
        pub_ids_raw = _strip(r.get("pub_ids"))
        pub_ids = _split_semicolon(pub_ids_raw)

        is_outstanding = False
        if _strip(r.get("is_outstanding")) != "":
            is_outstanding = _to_bool(r.get("is_outstanding"))

        outstanding_order = _to_int(
            r.get("outstanding_order"), ctx=f"people row {i} (id={pid}) outstanding_order"
        )

        if is_outstanding and role != "member":
            raise ImportErrorExit(
                f"people row {i} (id={pid}): is_outstanding=true requires role='member'"
            )

        person: Dict[str, Any] = {
            "id": pid,
            "role": role,
            "name": {"en": name_en, "zh": name_zh},
        }
        title_obj = {"en": title_en, "zh": title_zh}
        if title_en or title_zh:
            person["title"] = title_obj
        if join_year is not None:
            person["join_year"] = join_year
        if domain_list:
            person["domain"] = domain_list if len(domain_list) > 1 else domain_list[0]
        if photo:
            person["photo"] = photo
        bio_obj = {"en": bio_en, "zh": bio_zh}
        if bio_en or bio_zh:
            person["bio"] = bio_obj
        dest_obj = {"en": destination_en, "zh": destination_zh}
        if destination_en or destination_zh:
            person["destination"] = dest_obj
        if pub_ids:
            person["pub_ids"] = pub_ids
        if is_outstanding:
            person["is_outstanding"] = True
        if outstanding_order is not None:
            person["outstanding_order"] = outstanding_order

        people.append(person)
        id_to_name_en[pid] = name_en
        nn = _norm_name(name_en)
        if nn and nn not in name_norm_to_id:
            name_norm_to_id[nn] = pid

    # Validate outstanding_order uniqueness for those set.
    order_to_id: Dict[int, str] = {}
    for p in people:
        if not p.get("is_outstanding"):
            continue
        order = p.get("outstanding_order")
        if order is None:
            continue
        if order in order_to_id:
            raise ImportErrorExit(
                f"people: outstanding_order {order} duplicated for {order_to_id[order]!r} and {p['id']!r}"
            )
        order_to_id[order] = p["id"]

    # Fill missing outstanding_order deterministically:
    # remaining outstanding members are sorted by join_year desc then id asc,
    # assigned the smallest unused positive integers.
    outstanding_missing: List[Dict[str, Any]] = []
    for p in people:
        if p.get("is_outstanding") and p.get("outstanding_order") is None:
            outstanding_missing.append(p)

    if outstanding_missing:
        outstanding_missing.sort(
            key=lambda p: (
                -(int(p.get("join_year") or 0)),
                str(p.get("id") or ""),
            )
        )
        next_order = 1
        for p in outstanding_missing:
            while next_order in order_to_id:
                next_order += 1
            p["outstanding_order"] = next_order
            order_to_id[next_order] = p["id"]
            next_order += 1

    # Stable ordering in YAML
    people.sort(key=lambda p: str(p.get("id")))
    return people, PeopleIndex(id_to_name_en=id_to_name_en, name_norm_to_id=name_norm_to_id)


def _parse_publications(
    rows: List[Dict[str, Any]], people_index: PeopleIndex
) -> Tuple[List[Dict[str, Any]], Dict[str, List[Tuple[int, str]]]]:
    seen: set[str] = set()
    pubs: List[Dict[str, Any]] = []
    inferred: Dict[str, List[Tuple[int, str]]] = {}
    for i, r in enumerate(rows, start=2):
        pid = _strip(r.get("id"))
        if not pid:
            raise ImportErrorExit(f"publications row {i}: missing id")
        if pid in seen:
            raise ImportErrorExit(f"publications: duplicate id {pid!r}")
        seen.add(pid)

        title = _strip(r.get("title"))
        venue = _strip(r.get("venue"))
        year = _to_int(r.get("year"), ctx=f"publications row {i} (id={pid}) year")
        area = _strip(r.get("area"))
        if area not in ("EDA", "LCA"):
            raise ImportErrorExit(f"publications row {i} (id={pid}): invalid area {area!r}")
        authors_raw = _strip(r.get("authors"))
        link = _strip(r.get("link")) or "#"
        if not all([title, venue, year is not None, authors_raw]):
            raise ImportErrorExit(f"publications row {i} (id={pid}): missing required field(s)")

        authors_tokens = _split_semicolon(authors_raw)
        authors: List[str] = []
        for tok in authors_tokens:
            tok_s = _strip(tok)
            # record inferred relationships (by id token or by matching name_en)
            if tok_s in people_index.id_to_name_en:
                inferred.setdefault(tok_s, []).append((int(year), pid))
            else:
                nn = _norm_name(tok_s)
                mid = people_index.name_norm_to_id.get(nn)
                if mid:
                    inferred.setdefault(mid, []).append((int(year), pid))
            authors.append(people_index.id_to_name_en.get(tok_s, tok_s))

        pub: Dict[str, Any] = {
            "id": pid,
            "title": title,
            "authors": authors,
            "venue": venue,
            "year": year,
            "area": area,
            "link": link,
        }
        typ = _strip(r.get("type"))
        note = _strip(r.get("note"))
        if typ:
            pub["type"] = typ
        if note:
            pub["note"] = note

        pubs.append(pub)

    # Sort by year desc then id asc (matches existing display intent)
    pubs.sort(key=lambda p: (-int(p["year"]), str(p["id"])))
    return pubs, inferred


def _parse_projects(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen: set[str] = set()
    projects: List[Dict[str, Any]] = []
    for i, r in enumerate(rows, start=2):
        pid = _strip(r.get("id"))
        if not pid:
            raise ImportErrorExit(f"projects row {i}: missing id")
        if pid in seen:
            raise ImportErrorExit(f"projects: duplicate id {pid!r}")
        seen.add(pid)

        area = _strip(r.get("area"))
        if area not in ("EDA", "LCA"):
            raise ImportErrorExit(f"projects row {i} (id={pid}): invalid area {area!r}")

        status = _strip(r.get("status"))
        if status not in ("ongoing", "completed"):
            raise ImportErrorExit(f"projects row {i} (id={pid}): invalid status {status!r}")

        github = _strip(r.get("github")) or "#"
        lead_en = _strip(r.get("lead_en"))
        lead_zh = _strip(r.get("lead_zh"))
        title_en = _strip(r.get("title_en"))
        title_zh = _strip(r.get("title_zh"))
        summary_en = _strip(r.get("summary_en"))
        summary_zh = _strip(r.get("summary_zh"))
        start_year = _to_int(r.get("start_year"), ctx=f"projects row {i} (id={pid}) start_year")

        if not all([lead_en, lead_zh, title_en, title_zh, summary_en, summary_zh]):
            raise ImportErrorExit(f"projects row {i} (id={pid}): missing required bilingual fields")

        proj: Dict[str, Any] = {
            "id": pid,
            "area": area,
            "status": status,
            "github": github,
            "lead": {"zh": lead_zh, "en": lead_en},
            "title": {"zh": title_zh, "en": title_en},
            "summary": {"zh": summary_zh, "en": summary_en},
        }
        if start_year is not None:
            proj["start_year"] = start_year
        projects.append(proj)

    # Stable ordering
    projects.sort(key=lambda p: str(p["id"]))
    return projects


def _prune_content_people(root: Path) -> List[Path]:
    content_people = root / "content" / "people"
    deleted: List[Path] = []
    if not content_people.exists():
        return deleted
    for p in content_people.glob("*.md"):
        if p.name in ("_index.en.md", "_index.zh.md"):
            continue
        p.unlink()
        deleted.append(p)
    return deleted


def main(argv: List[str]) -> int:
    ap = argparse.ArgumentParser(description="Import JCIE site data from one XLSX.")
    ap.add_argument(
        "xlsx",
        nargs="?",
        default="xlsx/site.xlsx",
        help="Path to XLSX with sheets: people/publications/projects (default: xlsx/site.xlsx)",
    )
    ap.add_argument("--root", default=".", help="Repo root (default: current directory)")
    ap.add_argument(
        "--prune-content-people",
        action="store_true",
        help="Delete content/people/*.md except _index.{en,zh}.md",
    )
    args = ap.parse_args(argv)

    root = Path(args.root).resolve()
    xlsx = Path(args.xlsx).resolve()
    if not xlsx.exists():
        raise ImportErrorExit(f"XLSX not found: {xlsx}")

    wb = _load_workbook(xlsx)
    for sheet in ("people", "publications", "projects"):
        if sheet not in wb.sheetnames:
            raise ImportErrorExit(f"Missing required sheet: {sheet!r}. Found: {', '.join(wb.sheetnames)}")

    # people
    people_headers, people_rows = _read_sheet_rows(wb["people"])
    _require_columns(
        people_headers,
        ["id", "role", "name_en", "name_zh"],
        sheet="people",
    )
    people, people_index = _parse_people(people_rows)

    # publications
    pub_headers, pub_rows = _read_sheet_rows(wb["publications"])
    _require_columns(pub_headers, ["id", "title", "venue", "year", "area", "authors", "link"], sheet="publications")
    pubs, inferred_pub_ids = _parse_publications(pub_rows, people_index)

    # projects
    proj_headers, proj_rows = _read_sheet_rows(wb["projects"])
    _require_columns(
        proj_headers,
        [
            "id",
            "area",
            "status",
            "github",
            "lead_en",
            "lead_zh",
            "title_en",
            "title_zh",
            "summary_en",
            "summary_zh",
            "start_year",
        ],
        sheet="projects",
    )
    projs = _parse_projects(proj_rows)

    # Validate/fill people.pub_ids against publications.
    pub_id_set = {p["id"] for p in pubs}
    for p in people:
        pid = p["id"]
        pub_ids = p.get("pub_ids") or []
        if pub_ids:
            missing = [x for x in pub_ids if x not in pub_id_set]
            if missing:
                raise ImportErrorExit(f"people id={pid}: pub_ids references unknown publications: {', '.join(missing)}")
        else:
            inferred = inferred_pub_ids.get(pid) or []
            if inferred:
                inferred_sorted = sorted(inferred, key=lambda t: (-int(t[0]), t[1]))
                p["pub_ids"] = [pub_id for _, pub_id in inferred_sorted]

    # write yaml
    written: List[Path] = []
    people_yaml = {"people": people}
    pubs_yaml = {"publications": pubs}
    projs_yaml = {"projects": projs}

    written.append(root / "data" / "people.yaml")
    _write_text(written[-1], _yaml_write(people_yaml))
    written.append(root / "data" / "publications.yaml")
    _write_text(written[-1], _yaml_write(pubs_yaml))
    written.append(root / "data" / "projects.yaml")
    _write_text(written[-1], _yaml_write(projs_yaml))

    deleted: List[Path] = []
    if args.prune_content_people:
        deleted = _prune_content_people(root)

    print("Wrote:")
    for p in written:
        print(f"  - {p}")
    if deleted:
        print("Deleted:")
        for p in deleted:
            print(f"  - {p}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv[1:]))
    except ImportErrorExit as e:
        print(f"ERROR: {e}", file=sys.stderr)
        raise SystemExit(2)
