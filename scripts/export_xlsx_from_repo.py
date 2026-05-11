#!/usr/bin/env python3
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any, Dict, List, Tuple

import yaml


def _strip(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _load_yaml(path: Path) -> Dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def _norm_name(s: str) -> str:
    s = _strip(s).lower()
    # normalize common punctuation/spaces for matching
    for ch in [".", ",", "’", "'", "\"", "–", "-", "鈥?", "�"]:
        s = s.replace(ch, " ")
    s = " ".join(s.split())
    return s


def _split_semicolon(v: Any) -> List[str]:
    s = _strip(v)
    if not s:
        return []
    return [p.strip() for p in s.split(";") if p.strip()]


def _parse_front_matter(md_text: str) -> Dict[str, Any]:
    # Minimal Hugo front matter: --- yaml --- body
    parts = md_text.split("---")
    if len(parts) < 3:
        return {}
    fm = parts[1]
    try:
        data = yaml.safe_load(fm) or {}
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _load_alumni_destinations(root: Path) -> Dict[str, Dict[str, str]]:
    out: Dict[str, Dict[str, str]] = {}
    alumni_dir = root / "content" / "alumni"
    if not alumni_dir.exists():
        return out
    for path in alumni_dir.glob("*.md"):
        if path.name.startswith("_index."):
            continue
        # filename: <id>.<lang>.md
        stem = path.name[:-3]
        if stem.endswith(".en"):
            pid, lang = stem[:-3], "en"
        elif stem.endswith(".zh"):
            pid, lang = stem[:-3], "zh"
        else:
            continue
        fm = _parse_front_matter(path.read_text(encoding="utf-8", errors="replace"))
        dest = _strip(fm.get("destination"))
        if not dest:
            continue
        out.setdefault(pid, {"en": "", "zh": ""})
        out[pid][lang] = dest
    return out


def _build_people_sheet_rows(
    people_yaml: Dict[str, Any],
    *,
    destinations: Dict[str, Dict[str, str]],
    pub_ids_by_person: Dict[str, List[str]],
) -> List[Dict[str, Any]]:
    people = people_yaml.get("people") or []
    rows: List[Dict[str, Any]] = []
    for p in people:
        name = p.get("name") or {}
        title = p.get("title") or {}
        bio = p.get("bio") or {}
        dest = p.get("destination") or destinations.get(_strip(p.get("id")), {}) or {}
        pub_ids = p.get("pub_ids") or pub_ids_by_person.get(_strip(p.get("id")), []) or []
        domain = p.get("domain")
        if isinstance(domain, list):
            domain_str = ";".join([_strip(x) for x in domain if _strip(x)])
        else:
            domain_str = _strip(domain)
        rows.append(
            {
                "id": _strip(p.get("id")),
                "role": _strip(p.get("role")),
                "name_en": _strip(name.get("en")),
                "name_zh": _strip(name.get("zh")),
                "title_en": _strip(title.get("en")),
                "title_zh": _strip(title.get("zh")),
                "join_year": p.get("join_year") or "",
                "domain": domain_str,
                "photo": _strip(p.get("photo")),
                "bio_en": _strip(bio.get("en")),
                "bio_zh": _strip(bio.get("zh")),
                "destination_en": _strip(dest.get("en")),
                "destination_zh": _strip(dest.get("zh")),
                "pub_ids": ";".join([_strip(x) for x in pub_ids if _strip(x)]),
                "is_outstanding": "TRUE" if p.get("is_outstanding") else "",
                "outstanding_order": p.get("outstanding_order") or "",
            }
        )
    rows.sort(key=lambda r: r["id"])
    return rows


def _build_people_name_index(people_rows: List[Dict[str, Any]]) -> Dict[str, str]:
    # normalized name_en -> person id
    idx: Dict[str, str] = {}
    for r in people_rows:
        n = _norm_name(r.get("name_en") or "")
        if n:
            idx[n] = r["id"]
    return idx


def _build_person_pub_ids_from_publications(
    pub_rows: List[Dict[str, Any]], *, people_name_index: Dict[str, str]
) -> Dict[str, List[str]]:
    # person_id -> list of (year, pub_id)
    tmp: Dict[str, List[Tuple[int, str]]] = {}
    for r in pub_rows:
        pub_id = _strip(r.get("id"))
        year_val = r.get("year")
        try:
            year = int(year_val)
        except Exception:
            year = 0
        for tok in _split_semicolon(r.get("authors")):
            # authors column may contain person ids (preferred) or names
            if tok in people_name_index.values():
                tmp.setdefault(tok, []).append((year, pub_id))
            else:
                pid = people_name_index.get(_norm_name(tok))
                if pid:
                    tmp.setdefault(pid, []).append((year, pub_id))
    out: Dict[str, List[str]] = {}
    for pid, lst in tmp.items():
        lst_sorted = sorted(lst, key=lambda t: (-t[0], t[1]))
        out[pid] = [pub for _, pub in lst_sorted]
    return out


def _build_publications_sheet_rows(
    pubs_yaml: Dict[str, Any], *, people_name_index: Dict[str, str]
) -> List[Dict[str, Any]]:
    pubs = pubs_yaml.get("publications") or []
    rows: List[Dict[str, Any]] = []
    for p in pubs:
        authors = p.get("authors") or []
        author_tokens: List[str] = []
        for a in authors:
            aid = people_name_index.get(_norm_name(a))
            author_tokens.append(aid or _strip(a))
        rows.append(
            {
                "id": _strip(p.get("id")),
                "title": _strip(p.get("title")),
                "venue": _strip(p.get("venue")),
                "year": p.get("year") or "",
                "area": _strip(p.get("area")),
                "authors": ";".join([t for t in author_tokens if t]),
                "link": _strip(p.get("link")) or "#",
                "type": _strip(p.get("type")),
                "note": _strip(p.get("note")),
            }
        )
    # keep as-is ordering (already sorted in file), but stable fallback
    rows.sort(key=lambda r: (-(int(r["year"]) if str(r["year"]).isdigit() else 0), r["id"]))
    return rows


def _build_projects_sheet_rows(projects_yaml: Dict[str, Any]) -> List[Dict[str, Any]]:
    projs = projects_yaml.get("projects") or []
    rows: List[Dict[str, Any]] = []
    for p in projs:
        lead = p.get("lead") or {}
        title = p.get("title") or {}
        summary = p.get("summary") or {}
        rows.append(
            {
                "id": _strip(p.get("id")),
                "area": _strip(p.get("area")),
                "status": _strip(p.get("status")),
                "github": _strip(p.get("github")),
                "lead_en": _strip(lead.get("en")),
                "lead_zh": _strip(lead.get("zh")),
                "title_en": _strip(title.get("en")),
                "title_zh": _strip(title.get("zh")),
                "summary_en": _strip(summary.get("en")),
                "summary_zh": _strip(summary.get("zh")),
                "start_year": p.get("start_year") or "",
            }
        )
    rows.sort(key=lambda r: r["id"])
    return rows


def _write_sheet(ws, *, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(ord('A') + len(headers) - 1)}{len(rows) + 1}"


def main(argv: List[str]) -> int:
    ap = argparse.ArgumentParser(description="Export repo YAML data to a single XLSX template.")
    ap.add_argument("--root", default=".", help="Repo root (default: current directory)")
    ap.add_argument("--out", default="xlsx/site.xlsx", help="Output XLSX path (default: xlsx/site.xlsx)")
    args = ap.parse_args(argv)

    root = Path(args.root).resolve()
    out_path = (root / args.out).resolve()

    people_yaml = _load_yaml(root / "data" / "people.yaml")
    pubs_yaml = _load_yaml(root / "data" / "publications.yaml")
    projects_yaml = _load_yaml(root / "data" / "projects.yaml")

    # Build a name index from people.yaml (name_en -> id) for author/id substitution.
    people_tmp_rows = []
    for p in (people_yaml.get("people") or []):
        name = p.get("name") or {}
        people_tmp_rows.append({"id": _strip(p.get("id")), "name_en": _strip(name.get("en"))})
    people_name_index = _build_people_name_index(people_tmp_rows)

    pub_rows = _build_publications_sheet_rows(pubs_yaml, people_name_index=people_name_index)
    proj_rows = _build_projects_sheet_rows(projects_yaml)
    pub_ids_by_person = _build_person_pub_ids_from_publications(pub_rows, people_name_index=people_name_index)
    destinations = _load_alumni_destinations(root)

    people_rows = _build_people_sheet_rows(
        people_yaml,
        destinations=destinations,
        pub_ids_by_person=pub_ids_by_person,
    )

    try:
        import openpyxl  # type: ignore
    except Exception as e:
        raise SystemExit(
            "openpyxl is required to write XLSX. Install with: python -m pip install openpyxl"
        ) from e

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    ws_people = wb.create_sheet("people")
    people_headers = [
        "id",
        "role",
        "name_en",
        "name_zh",
        "title_en",
        "title_zh",
        "join_year",
        "domain",
        "photo",
        "bio_en",
        "bio_zh",
        "destination_en",
        "destination_zh",
        "pub_ids",
        "is_outstanding",
        "outstanding_order",
    ]
    _write_sheet(ws_people, headers=people_headers, rows=people_rows)

    ws_pubs = wb.create_sheet("publications")
    pub_headers = ["id", "title", "venue", "year", "area", "authors", "link", "type", "note"]
    _write_sheet(ws_pubs, headers=pub_headers, rows=pub_rows)

    ws_projects = wb.create_sheet("projects")
    proj_headers = [
        "id",
        "area",
        "status",
        "github",
        "lead_en",
        "lead_zh",
        "title_en",
        "title_zh",
        "summary_en",
        "summary_zh",
        "start_year",
    ]
    _write_sheet(ws_projects, headers=proj_headers, rows=proj_rows)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
